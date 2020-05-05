# -*- coding: utf-8 -*-
"""
:::::::::::::::::::::::::::::::::::::  C-I-C-A-T  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE

This software (or technical data) was produced for the U. S. Government under contract TVD, and is 
subject to the Rights in Data-General Clause 52.227-14 (MAY 2014) or (DEC 2007).

The following copyright notice may be affixed after receipt of written approval from the Contracting Officer.
Please contact the Contracts Management Office for assistance with obtaining approval or identifying the correct clause.
If the contract has Clause 52.227-14, Alt. IV, written approval is not required and the below copyright notice may be affixed.

(c) 2020 The MITRE Corporation. All Rights Reserved.


spreadout.py - Python module implementing export spreadsheet generation for Tom Sawyer visualization

:::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""

#import sys
from openpyxl import Workbook
from collections import defaultdict

m_exportSheets = ['SCENARIO_DESC', 'SCEN_ENTRY_POINT', 'THREAT_ACTOR_DESC', 'ASSET_TTP_MAP','TTP_MITIGATION_MAP', 'TTP_DESC', 'MITIGATION_DESC', 'ATTACK_PATH'] 

m_HEADER_tags = defaultdict(list )

m_HEADER_tags[0] = ['SCN_ID', 'SHORT_NAME', 'NAME', 'IMPACT_SCORE', 'INTENT', 'DESCRIPTION', 'EXTERNAL_INSIDER_EXTASINSIDER', 'DETAIL', 'ACTOR', 'TTP_PATTERN', 'ASSET_NAME' ]
m_HEADER_tags[1] = ['SCN_ID', 'ENTRY_ASSET_NAME' ]
m_HEADER_tags[2] = ['GROUP_ID', 'ACTOR', 'ALIASES', 'SOPHISTICATION']
m_HEADER_tags[3] = ['SCN_ID', 'ENTRY_ASSET_NAME', 'ASSET_ID', 'TTP_ID']
m_HEADER_tags[4] = ['TTP_ID', 'MITIGATION_ID']
m_HEADER_tags[5] = ['TTP_ID', 'NAME', 'DESCRIPTION', 'URL', 'TACTICS_LIST']
m_HEADER_tags[6] = ['MITIGATION_ID', 'NAME', 'DESCRIPTION', 'URL']
m_HEADER_tags[7] = ['SCENARIO_ID', 'ENTRY_ASSET_NAME', 'FROM_ASSET_ID', 'TO_ASSET_ID']

def list2string(plist, separator):
    ret = 'undefined'
    if plist:
       ret = str(plist[0])
       for j in range (1, len(plist)):
           ret = ret + separator + str(plist[j])
    return ret

def getTTPlist (tracedata):
    ttplist = []
    for k in tracedata.keys():
        curdict = tracedata[k]
        curtrace = curdict[0]
        tlist = curtrace['ttps']   
        for x in tlist:
           if len(x) > 0:
               for xx in x:
                  ttplist.append(xx)

    tset = set (ttplist)  # dump list into set to elimnate duplicates
    ret = list (tset)     # dump back into  list

    return ret

def getMITsforTTP (dataset, ttpID ):
    for t in dataset['ATT&CK']:
        ret = []
        if t.getTECHID() == ttpID:
            clist = t.getCOA()
            for c in clist:
               ret.append(c[0].getTECHID() )
            return ret
 

def getMITsforTTPList (dataset, ttplist):
    clist = []
    for x in ttplist:
        moremits = getMITsforTTP(dataset, x)
        if moremits:
          clist = clist + moremits

    uset = set (clist)
    ret = list (uset)
    
    return ret

def getFromToList(path):
    ret = []
    idx = 0
    while idx < len(path)-1 :
        ret.append([path[idx], path[idx+1]])
        idx = idx + 1
    
    return ret
          
def saveSheetHeaders(sheetname, sheet):    
    sheetid = m_exportSheets.index(sheetname)
    colindx = 1
    for k in m_HEADER_tags[sheetid]:
        sheet.cell (1, colindx, k)
        colindx = colindx + 1
        
def saveSheetData (sheetname, sheet, tracedata, dataset):

    if sheetname == 'SCENARIO_DESC':
        row = 2
        for s in dataset['SCENARIO']:
           sheet.cell(row, 1, s.getID() ) 
           sheet.cell(row, 2, s.getShortName() ) 
           sheet.cell(row, 3, s.getName())
           sheet.cell(row, 4, '#IMPACT SCORE')
           sheet.cell(row, 5, '#INTENT')
           sheet.cell(row, 6, s.getDesc() )
           sheet.cell(row, 7, '#EXTERNAL_INSIDER_EXTRAINSIDER')           
           sheet.cell(row, 8, '#DETAIL' )
           sheet.cell(row, 9, s.getActorID() )
           sheet.cell(row, 10, s.getIntendedEffect () )
           sheet.cell(row, 11, s.getTargetID() )     
           row = row + 1    

    elif sheetname == 'SCEN_ENTRY_POINT':     
        row = 2
        for k in tracedata.keys():
            values = k.split ('EP', maxsplit=1)
            sheet.cell (row, 1, values[0] )
            sheet.cell (row, 2, values[1] )
            row = row + 1

    elif sheetname == 'THREAT_ACTOR_DESC':
        row = 2
        for s in dataset['SCENARIO']:
            actorID = s.getActorID()
            for a in dataset['ATKGROUPS']:
                if a.getName() == actorID:
                    sheet.cell (row, 1, a.getGroupID() )
                    sheet.cell (row, 2, a.getName () )
                    sheet.cell (row, 3, list2string (a.getAliases(), ', ' ) )
                    sheet.cell (row, 4, a.getSophisticationLevel() )
                    break            
            row = row + 1

    elif sheetname == 'ASSET_TTP_MAP':     
        row = 2
        for k in tracedata.keys():
            values = k.split ('EP', maxsplit=1)
            curdict = tracedata[k]
            curtrace = curdict[0]
            clist = curtrace['path']
            tlist = curtrace['ttps']
            
            idx = 0  #index controls mapping of ttps to  assets in component list
            for c in clist:
               for t in tlist[idx]:
                  sheet.cell (row, 1, values[0] )
                  sheet.cell (row, 2, values[1] )
                  sheet.cell (row, 3, c)
                  sheet.cell (row, 4, t)
                  row = row + 1
               idx = idx + 1   

    elif sheetname == 'TTP_DESC':       
        ttplist = getTTPlist (tracedata)
        row = 2
        for ttp in ttplist:
                for d in dataset['ATT&CK']:
                    if d.getTECHID() == ttp:
                        sheet.cell (row, 1, d.getTECHID() )
                        sheet.cell (row, 2, d.getName() )
                        sheet.cell (row, 3, d.getDesc() )
                        sheet.cell (row, 4, d.getURL() )
                        sheet.cell (row, 5, list2string (d.getTactic(), ', ' ))
                        break
                row = row + 1

    elif sheetname == 'TTP_MITIGATION_MAP':
        ttplist = getTTPlist (tracedata)
        row = 2
        for ttp in ttplist:
            clist = getMITsforTTP (dataset, ttp)
            if clist:
               for c in clist:
                   sheet.cell (row, 1, ttp)
                   sheet.cell (row, 2, c)
                   row = row + 1
                
    elif sheetname == 'MITIGATION_DESC':
        ttplist = getTTPlist (tracedata)
        clist = getMITsforTTPList (dataset, ttplist)
        row = 2
        for c in clist: 
            for q in dataset ['ATKMITIGATION']:
                if q.getTECHID() == c:
                    sheet.cell (row, 1, q.getTECHID() )
                    sheet.cell (row, 2, q.getName() )
                    sheet.cell (row, 3, q.getDesc() )
                    sheet.cell (row, 4, q.getURL() )
                    break
            row = row + 1

    elif sheetname == 'ATTACK_PATH':
        row = 2
        for k in tracedata.keys():
            values = k.split ('EP', maxsplit=1)
            curdict = tracedata[k]
            curtrace = curdict[0]
            clist = curtrace['path']   
            fromtolist = getFromToList (clist) 
            if fromtolist:
                for p in fromtolist:
                  sheet.cell (row, 1, values[0] )
                  sheet.cell (row, 2, values[1] )
                  sheet.cell (row, 3, p[0] )
                  sheet.cell (row, 4, p[1] )
                  row = row + 1
                                  
           
def ExportSheet (wb, sheetname, tracedata, dataset):   
    sheet = wb.create_sheet (title=sheetname )
    saveSheetHeaders(sheetname, sheet)
    saveSheetData (sheetname, sheet, tracedata, dataset )
               
def ExportData (fname, tracedata, dataset ):
    wb = Workbook()     
    for sheetname in m_exportSheets:
        ExportSheet (wb, sheetname, tracedata, dataset)
    wb.save(filename = fname )   


# main entry point
if ( __name__ == "__main__"):  
   print ('No unit test currently supported.')

       





        